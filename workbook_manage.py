from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import Reference, LineChart
import os
import control_variables as cv


class Workbook_Manage:
    __workbook = None
    __worksheet = None
    __workbook_path = ''

    def __init__(self, ws_n):
        self.__workbook_path = self.__get_workbook_path()
        self.__open_workbook()
        if ws_n in self.__workbook.sheetnames:
            self.__worksheet = self.__workbook.get_sheet_by_name(ws_n)
            self.write_array([])
            self.write_array([])
        else:
            self.__worksheet = self.__workbook.create_sheet(ws_n)

    def __del__(self):
        self.__close_workbook()

    def write_dataframe(self, df):
        for row in dataframe_to_rows(df, False):
            self.write_array(row)

    def write_string(self, str):
        self.write_array([str])

    def write_array(self, arr):
        self.__worksheet.append(arr)

    def write_line_chart(self, title, data_fields, axis_fields, pos):
        chart = LineChart()
        data = Reference(self.__worksheet, min_row=data_fields[0], max_row=data_fields[1], min_col=data_fields[2], max_col=data_fields[3])
        chart.add_data(data, titles_from_data=True)
        categories = Reference(self.__worksheet, min_row=axis_fields[0], max_row=axis_fields[1], min_col=axis_fields[2])
        chart.set_categories(categories)
        chart.title = title
        self.__worksheet.add_chart(chart, pos)

    def __get_workbook_path(self):
        return cv.common_fname + '.xlsx'

    def __open_workbook(self):
        if os.path.exists(self.__workbook_path):
            self.__workbook = load_workbook(self.__workbook_path)
        else:
            self.__workbook = Workbook()

    def __close_workbook(self):
        self.__workbook.save(self.__workbook_path)

    
